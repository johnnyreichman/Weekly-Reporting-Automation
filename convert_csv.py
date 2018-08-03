#This program converts xlsx/xlsm files into csv format

import openpyxl
from openpyxl import load_workbook
import csv


def exportCSV(input_file, output_file):
	print("Converting to csv")
	row_count = 0
	wb = openpyxl.load_workbook(input_file, read_only = True)
	sh = wb.active
	with open(output_file, 'w', newline = "") as f:  # open('test.csv', 'w', newline="") for python 3
	    c = csv.writer(f)
	    for r in sh.rows:
	    	row_count += 1
	    	print(row_count, end = "\r")
	    	c.writerow([cell.value for cell in r]) #maybe fix this so dates get cut off?


#This one ensures that only the date is written into the csv file, and not the time as well(00:00:00)
def exportCSV_with_dates(input_file, output_file, date_column):
	print("Converting to csv")
	row_count = 0
	wb = openpyxl.load_workbook(input_file, read_only = True)
	sh = wb.active
	with open(output_file, 'w', newline = "") as f:  # open('test.csv', 'w', newline="") for python 3
	    c = csv.writer(f)
	    for r in sh.rows:
	    	row_count += 1
	    	print(row_count, end = "\r")
	    	cell_count = 0
	    	app_row = []
	    	for cell in r:
	    		cell_count += 1
	    		if cell_count == date_column:
	    			app_row.append(str(cell.value)[:10])
	    		else:
	    			app_row.append(cell.value)
	    	c.writerow(app_row)




#This programn runs multiple macros on multiple (large) input files
#What it does:
#	creates a new file and places a macro in it
#	copies active sheet from the input and puts it into the new file
#	runs macro on the new file
#	saves the new file after the macro has run


#various imports used for different functions
import os, os.path
import win32com.client
from openpyxl import load_workbook
import xlsxwriter
from win32com.client import Dispatch


#Why the code is so long:
#	Built-in macro running function only runs macro in the file it resides: 
#	solution--> copy input file into macro file then run macro. Takes longer but only O(n)


#Takes in inout_file.xlsx, macro.bin, output_file creates input_file.xlsm
def addMacro(input_file, macro_bin_name, output_file = "default_output"):
	save_name = output_file + ".xlsm"
	wb2 = xlsxwriter.Workbook(save_name)
	wb2.add_vba_project(macro_bin_name)
	wb2.close()

#Takes in input_file, macro_name, macro_destination and runs the macro attached
def runMacro(input_file, macro_name, output_file = "default_output"):

	#copies contents of input_file into macro_file and saves it as output_file
	xl = Dispatch("Excel.Application")
	xl.Visible = False  # You can remove this line if you don't want the Excel application to be visible
	wb1 = xl.Workbooks.Open(os.path.join(os.getcwd(), input_file))
	wb2 = xl.Workbooks.Open(os.path.join(os.getcwd(), output_file + ".xlsm"))
	ws1 = wb1.Worksheets(1)
	ws1.Copy(Before=wb2.Worksheets(1))
	
	print("Running Macro")
	xl.Run(macro_name)
	print("Saving macro output")
	wb2.Close(SaveChanges=True)
	wb1.Close()
	xl.Quit()


	#UNCOMMENT THIS CHUNK IF YOU WANT THE OOUTPUT IN XLSX FORMAT TOO
	#Reopens output_file after macro has run, but opens it 
	#without the macro included so it can be converted to .xlsx
	# print("Converting macro output to xlsx")
	# wb = load_workbook(filename=output_file + ".xlsm", keep_vba=False)
	# wb.save(output_file + ".xlsx")
	# wb.close()

	print("Macro ran successfully!")


